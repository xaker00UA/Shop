from typing import Any
from aiogram import F, Router
from aiogram.types import CallbackQuery, Message
from aiogram.filters import Command, Filter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State,StatesGroup
from core import keyboard
from core.database import Database
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
manager = Router()


class Order(StatesGroup):
    id=State()
    message=State()







class ManagerProtect(Filter):
    async def __call__(self, event):
        manager = await Database.Manager().get_all()
        manager = [d["id"] for d in manager]
        if isinstance(event,Message):
            return event.from_user.id in manager
        if isinstance(event,CallbackQuery):
            return event.from_user.id in manager
        


@manager.callback_query(ManagerProtect(), Command("manager"))
async def manager_menu(message:Message):
    await message.answer("Главное меню",reply_markup=keyboard.manager)

# Создаем таблицу
@manager.callback_query(ManagerProtect(), F.data == "create_table")
async def create_table(callback:CallbackQuery):
    callback.answer("")
    orders = await Database.Order().get_order_open()
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"
    keys = set()
    for order in orders:
        keys.update(order.keys())
    keys = list(keys)
    for col, title in enumerate(keys,1):
        ws.cell(row=1,column=col,value=title)
    row=2   
    for order in orders:
        for key, value, in order.items():
            ws.cell(row=row,column=keys.index(key)+1,value=value)
            row+=1
    product=await Database.Product.get_all()
    index=[keys.index(i) for i in product if i in keys]
    for col in index:
        formula=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{(row-1)})"
        ws.cell(row=row,column=col,value=formula)
    table=BytesIO
    wb.save(table)
    await callback.message.answer(document=BytesIO(table.getvalue()),filename="data.xlsx",replay_markup=keyboard.manager)
    table.close()




# Изменяем статус заказа
@manager.callback_query(ManagerProtect(), F.data == "change_order")
async def change_order(callback:CallbackQuery,state:FSMContext):
    await callback.answer("")
    await state.set_state(Order.id)
    await callback.message.answer("Введите номер или id заказа")
# Получаем id 
@manager.message(ManagerProtect(), Order.id)
async def search_order(message:Message,state:FSMContext):
    order = await Database.Order().get_order(message.text)
    if len(order) == 1:
        order = order[0]
        order["Статус"]="Закрыт"
        await Database.Order().add(order)
        await state.clear()
        await message.answer("Заказ закрыт",reply_markup=keyboard.manager)
        await state.clear()
    elif len(order) >= 1:
        await message.answer("Введите id-заказа полностью")
    else:
        await message.answer("Заказ не найден",reply_markup=keyboard.manager)
        await state.clear()
# Удаление закрытых заказов
@manager.callback_query(ManagerProtect(), F.data == "delete_order_close")
async def delete_order_close(callback:CallbackQuery):
    await callback.answer("")
    count = await Database.Order().delete_order_close()
    await callback.answer(f"Удалено {count} заказов")





# Рассылка сообщений
@manager.callback_query(ManagerProtect(), F.data == "write_order_open")
async def message(callback:CallbackQuery,state:FSMContext):
    await callback.answer("")
    state.set_state(Order.message)
    await callback.message.answer("Введите сообщение")

@manager.message(ManagerProtect(), Order.message)
async def send_message(message:Message,state:FSMContext):
    open = await Database.Order().get_order_open()
    open=[int(d["id"])for d in open]
    for user_id in open:
        await message.bot.send_message(user_id,message.text)
    await message.answer("Рассылка завершена")
    state.clear()


@manager.callback_query(ManagerProtect(), F.data == "")
async def protect(callback:CallbackQuery):
    pass

